"""重建商品档案（两级商品模型版）：
- 库存商品（大类·真实库存）：柠檬云商品 + 纸箱(包材) + 人工打包费 + 快递费 + 从七月SKU自动提炼的大类
- 订单商品（小类·出库销售）：七月干货/蔬菜 SKU，关联到库存商品，倍数 = 每单消耗的库存默认单位数
- 规则全部来自 product_rules.json（代码层面，方便维护）
"""
import http.cookiejar
import json
import re
import urllib.request
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

from app.database import SessionLocal
from app.models import CodeMapping, Product

BASE = "http://127.0.0.1:8000"
LEMON = "柠檬云商品导入模板.xlsx"
ROOT = Path(__file__).resolve().parent
CFG = json.loads((ROOT / "product_rules.json").read_text(encoding="utf-8"))
OPENER = None

SRC_FILES = CFG["source_files"]
SKIP = set(CFG["exclude_names"])
SKIP_PREFIX = tuple(CFG["exclude_prefixes"])
STRIP_RE = re.compile(CFG["strip_suffix_pattern"])
UNIT = CFG["unit_rules"]
BOX_NAMING = CFG["box_naming"]
BOX_PRICE = CFG["box_price"]
CATS = CFG["categories"]
LABOR = CFG["labor"]
SHIP = CFG["shipping"]
OVERRIDES = CFG.get("product_overrides", {})

# 规格后缀（用于从订单商品名提炼大类与倍数）
SPEC_SUFFIX_RE = re.compile(r"([\d.]+)\s*(个|斤|公斤|千克|克|g|kg|袋|包|盒|箱|件|份)\s*(装)?$")
GRAM = {"克": 1.0, "g": 1.0, "斤": 500.0, "公斤": 1000.0, "千克": 1000.0, "kg": 1000.0}


# ---------- HTTP（柠檬云导入） ----------
def login():
    global OPENER
    cj = http.cookiejar.CookieJar()
    OPENER = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))
    req = urllib.request.Request(BASE + "/api/auth/login",
                                 data=json.dumps({"username": "admin1", "password": "admin1"}).encode(),
                                 headers={"Content-Type": "application/json"})
    with OPENER.open(req) as r:
        return json.loads(r.read().decode())


def upload(path):
    with open(path, "rb") as f:
        data = f.read()
    fname = path.split("\\")[-1]
    boundary = "----bnd1234"
    body = (f"--{boundary}\r\nContent-Disposition: form-data; name=\"file\"; filename=\"{fname}\"\r\n"
            f"Content-Type: application/octet-stream\r\n\r\n").encode() + data + f"\r\n--{boundary}--\r\n".encode()
    req = urllib.request.Request(BASE + "/api/import/products", data=body, method="POST",
                                 headers={"Content-Type": f"multipart/form-data; boundary={boundary}"})
    with OPENER.open(req) as r:
        return json.loads(r.read().decode())


# ---------- 规则（来自配置） ----------
def clean_name(raw):
    return STRIP_RE.sub("", str(raw).strip()).strip()


def parse_boxes(cell):
    out = defaultdict(int)
    if not cell:
        return out
    for seg in str(cell).split("+"):
        seg = seg.strip()
        if not seg:
            continue
        parts = [p.strip() for p in seg.split("*")]
        model = parts[0]
        if not model:
            continue
        if len(parts) >= 2 and parts[-1].isdigit():
            count = int(parts[-1])
        else:
            count = len(parts)
        out[model] += count
    return out


def box_product_name(model):
    if re.fullmatch(r"\d+号", model):
        return f"{BOX_NAMING['numbered_prefix']}{model}"
    return model if model.endswith("箱") else f"{model}{BOX_NAMING['non_numbered_suffix']}"


def strip_spec(name):
    """从订单商品名剥离尾部规格 → (大类名, 数量, 单位)。如 '佛手柑大果2个'→('佛手柑大果',2,'个')"""
    m = SPEC_SUFFIX_RE.search(str(name))
    if m:
        unit = {"g": "克", "kg": "千克"}.get(m.group(2), m.group(2))
        base = name[: m.start()].strip()
        return base, float(m.group(1)), unit
    return None


def convert_qty(qty, from_unit, to_unit):
    if from_unit == to_unit:
        return qty
    if from_unit in GRAM and to_unit in GRAM:
        return qty * GRAM[from_unit] / GRAM[to_unit]
    return None  # 重量↔计数不可换算，跳过关联


def collect_july():
    """返回 (商品配置, 纸箱价格列表)。"""
    prod_cfg = defaultdict(lambda: {"cats": set(), "confs": defaultdict(int)})
    box_prices = defaultdict(list)
    for cat_label, path in SRC_FILES.items():
        wb = load_workbook(ROOT / path, data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=3, values_only=True):
                vals = list(row) + [None] * 10
                raw = vals[0]
                if not raw:
                    continue
                name = clean_name(raw)
                if not name or name in SKIP or name.startswith(SKIP_PREFIX):
                    continue
                try:
                    qty = int(float(vals[1])) if vals[1] else 1
                except (ValueError, TypeError):
                    qty = 1
                boxes = parse_boxes(vals[2])
                price = vals[4]
                labor = vals[6]
                bs = str(vals[2]).strip() if vals[2] else ""
                if bs and "*" not in bs and "+" not in bs and price and BOX_PRICE["single_box_only"]:
                    try:
                        box_prices[bs].append(float(price))
                    except (ValueError, TypeError):
                        pass
                if labor is not None:
                    try:
                        labor = round(float(labor), 2)
                    except (ValueError, TypeError):
                        labor = 0
                else:
                    labor = 0
                key = (tuple(sorted(boxes.items())), labor)
                prod_cfg[name]["cats"].add(cat_label)
                prod_cfg[name]["confs"][key] += qty
    return prod_cfg, box_prices


def best_config(conf_counts):
    return max(conf_counts.items(), key=lambda kv: kv[1])[0]


def box_unit_cost(model, prices):
    if model in BOX_PRICE.get("overrides", {}):
        return float(BOX_PRICE["overrides"][model])
    if not prices:
        return 0
    if BOX_PRICE.get("mode") == "mode":
        return round(Counter(prices).most_common(1)[0][0], 4)
    return round(sum(prices) / len(prices), 4)


def main():
    login()
    print("== 1) 导入柠檬云商品（库存商品/大类）==")
    r = upload(LEMON)
    print(f"   创建 {r['created']}，跳过 {r.get('skipped', 0)}")

    prod_cfg, box_prices = collect_july()
    print(f"\n== 2) 七月表解析 ==")
    print(f"   订单商品 {len(prod_cfg)} 个，纸箱型号 {len(box_prices)} 种")

    db = SessionLocal()
    try:
        existing = {p.name: p for p in db.query(Product).all()}

        # 库存商品目录（名称 → Product）
        stock_by_name = {}

        # 纸箱（包材，库存商品）
        box_ids = {}
        for model, prices in sorted(box_prices.items()):
            cost = box_unit_cost(model, prices)
            name = box_product_name(model)
            p = existing.get(name)
            if not p:
                p = Product(name=name, category=CATS["box"], product_type="stock",
                            base_unit="个", default_unit="个", conversions={"个": 1},
                            unit_cost=cost, spec=f"纸箱 {model}（参考 {cost:.2f} 元/个）", is_active=True)
                db.add(p)
                db.flush()
            box_ids[model] = p.id
            stock_by_name[name] = p
        print(f"   纸箱（包材/库存）：{len(box_ids)} 个")

        # 人工 / 快递（库存型参考商品）
        for pname, pcat, pspec in [
            ("人工打包费", CATS["labor"], "人工打包服务（可调整成本）"),
            (SHIP["product_name"], CATS["shipping"], "快递物流费用（可自行调整）"),
        ]:
            if pname not in existing:
                p = Product(name=pname, category=pcat, product_type="stock",
                            base_unit="单", default_unit="单", conversions={"单": 1},
                            spec=pspec, is_active=True)
                db.add(p)
                db.flush()
                stock_by_name[pname] = p
            else:
                stock_by_name[pname] = existing[pname]

        # 柠檬云商品 → 库存商品（已由导入创建）
        lemon_names = [p.name for p in db.query(Product).filter(Product.category == CATS["lemon"]).all()]
        for ln in lemon_names:
            stock_by_name[ln] = existing[ln]

        # 订单商品（七月 SKU），自动关联库存商品 + 倍数
        created_order = 0
        created_stock = 0
        created_labor = 0
        linked = 0
        unlinked = []
        for name, cfg_info in sorted(prod_cfg.items()):
            if name in existing:
                continue
            (boxes_tuple, labor) = best_config(cfg_info["confs"])
            cat = "干货" if "干货" in cfg_info["cats"] else "蔬菜"

            # 关联库存商品 + 倍数
            stock_id, multiplier = None, 1.0
            sp = strip_spec(name)
            if sp:
                base, qty, unit = sp
                target = stock_by_name.get(base)
                if not target:
                    # 尝试柠檬云大类子串匹配（如 京鲜生七彩花生2斤 → 七彩花生）
                    hits = [s for s in lemon_names if s in name]
                    if hits:
                        best = max(hits, key=len)
                        target = stock_by_name[best]
                        mul = convert_qty(qty, unit, target.default_unit or target.base_unit)
                        if mul is not None:
                            stock_id, multiplier = target.id, mul
                else:
                    mul = convert_qty(qty, unit, target.default_unit or target.base_unit)
                    if mul is not None:
                        stock_id, multiplier = target.id, mul
                if not target and stock_id is None:
                    # 新建大类库存商品
                    is_weight = unit in GRAM
                    du = unit if unit in ("斤", "个", "袋", "包", "盒", "箱", "件", "份") else ("克" if is_weight else "个")
                    nstock = Product(name=base, category=cat, product_type="stock",
                                     base_unit="克" if is_weight else "个",
                                     default_unit=du,
                                     conversions={"克": 1, "斤": 500, "公斤": 1000, "千克": 1000} if is_weight else {du: 1},
                                     is_active=True)
                    db.add(nstock)
                    db.flush()
                    stock_by_name[base] = nstock
                    stock_id, multiplier = nstock.id, qty
                    created_stock += 1
            if stock_id:
                linked += 1
            else:
                unlinked.append(name)

            # 独立打包费商品（库存型参考）
            labor_id = None
            labor_name = LABOR["product_name_template"].format(name=name)
            if labor and labor > 0:
                lp = stock_by_name.get(labor_name)
                if not lp:
                    lp = Product(name=labor_name, category=CATS["labor"], product_type="stock",
                                 base_unit=LABOR["unit"], default_unit=LABOR["unit"],
                                 conversions={LABOR["unit"]: 1}, unit_cost=labor,
                                 spec=f"{name} 打包费（{labor:.2f} 元/单）", is_active=True)
                    db.add(lp)
                    db.flush()
                    stock_by_name[labor_name] = lp
                    created_labor += 1
                labor_id = lp.id

            # 关联结算清单：纸箱 + 独立打包费
            pack_items = []
            if boxes_tuple:
                for model, c in boxes_tuple:
                    if model in box_ids:
                        pack_items.append({"product_id": box_ids[model], "quantity": c, "unit": "个"})
            if labor_id:
                pack_items.append({"product_id": labor_id, "quantity": LABOR["consume_per_order"], "unit": LABOR["unit"]})

            db.add(Product(
                name=name, category=cat, product_type="order",
                base_unit="单", default_unit="单", spec="",
                conversions={"单": 1},
                pack_items=pack_items, pack_fee=0,
                stock_product_id=stock_id, multiplier=multiplier,
                is_active=True,
            ))
            created_order += 1
        db.commit()

        # 编码关联初始化
        seeded_map = 0
        for cm in CFG.get("code_mappings", []):
            ext = str(cm.get("external_code", "")).strip()
            pname = str(cm.get("product_name", "")).strip()
            if not ext or not pname:
                continue
            p = db.query(Product).filter(Product.name == pname).first()
            if not p:
                print(f"   [跳过编码关联] 商品「{pname}」不存在")
                continue
            m = db.scalar(select(CodeMapping).where(
                CodeMapping.source == cm.get("source", "jushuitan"), CodeMapping.external_code == ext))
            if m:
                m.product_id = p.id
                m.auto_score = float(cm.get("score", 1.0))
            else:
                db.add(CodeMapping(source=cm.get("source", "jushuitan"), external_code=ext,
                                   external_name=ext, product_id=p.id,
                                   auto_score=float(cm.get("score", 1.0))))
            seeded_map += 1
        db.commit()
        print(f"   编码关联初始化 {seeded_map} 条")

        total = db.query(Product).count()
        types = Counter(db.query(Product.product_type).all())
        cats = Counter(db.query(Product.category).all())
        print(f"\n== 完成：商品总数 {total}，类型分布 {dict(types)} ==")
        print(f"   新增订单商品 {created_order} 个（其中 {linked} 个已自动关联库存商品），提炼库存大类 {created_stock} 个，打包费 {created_labor} 个")
        if unlinked:
            print(f"   未自动关联（需手动设置）{len(unlinked)} 个：{unlinked[:20]}")

        demo = db.query(Product).filter(Product.name == "佛手柑中果1个").first()
        if demo:
            sp = db.get(Product, demo.stock_product_id) if demo.stock_product_id else None
            packs = [(db.get(Product, it["product_id"]).name, it["quantity"]) for it in demo.pack_items]
            print(f"\n== 示例验证：佛手柑中果1个 ==\n   订单商品 → 库存商品：{sp.name if sp else '无'} ×{demo.multiplier}\n   关联结算清单: {packs}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
