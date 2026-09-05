"""批量导入：模板下载、商品/入库/出库导入、聚水潭出库单解析与商品编码关联。"""
import difflib
import io
import re
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from sqlalchemy import or_, select
from sqlalchemy.orm import Session

from ..auth import get_current_user
from ..database import get_db
from ..models import CodeMapping, Product, User
from ..services import (
    create_inbound,
    create_outbound,
    default_conversions,
    fmt_qty,
    resolve_product,
    unit_to_base,
)

router = APIRouter(prefix="/api", tags=["import"])

# 导入模版文件目录（后端 docs 目录）
DOCS_DIR = Path(__file__).resolve().parent.parent.parent / "backend" / "docs"

WEIGHT_UNITS = {"克": 1.0, "g": 1.0, "斤": 500.0, "公斤": 1000.0, "千克": 1000.0, "kg": 1000.0}

# ---------------- 表头别名 ----------------
PRODUCT_ALIASES = {
    "code": ["商品编码", "*商品编码", "编码"],
    "name": ["商品名称", "*商品名称", "名称", "商品名"],
    "category": ["商品类别", "商品分类", "分类", "类别", "类目"],
    "spec": ["规格型号", "规格", "型号"],
    "unit": ["单位", "*单位", "基本单位"],
    "base_unit": ["基础单位"],
    "conversions": ["单位换算"],
    "sale_price": ["默认售价", "销售价", "售价"],
    "pack_fee": ["打包费", "包装费"],
    "pack_items": ["关联商品清单", "关联商品"],
}
INBOUND_ALIASES = {
    "product": ["商品编码或名称", "商品编码", "商品名称", "商品", "名称"],
    "unit": ["进货单位", "单位", "*单位"],
    "quantity": ["数量", "*数量"],
    "unit_price": ["单价", "进货单价", "*单价"],
    "supplier": ["供应商"],
    "date": ["日期", "入库日期"],
    "operator": ["操作员"],
    "remark": ["备注"],
}
OUTBOUND_ALIASES = {
    "doc_no": ["单号", "订单号"],
    "date": ["日期", "出库日期"],
    "customer": ["客户"],
    "product": ["商品编码或名称", "商品编码", "商品名称", "商品"],
    "unit": ["销售单位", "单位", "出库单位"],
    "quantity": ["数量"],
    "unit_price": ["单价", "售价", "出库单价"],
    "pack_fee": ["打包费"],
    "operator": ["操作员"],
    "remark": ["备注"],
}
JUSHUITAN_COLS = {
    "doc_no": ["出库单号", "出仓单号"],
    "date": ["出库日期"],
    "status": ["状态"],
    "name": ["商品名称"],
    "amount": ["卖家实收", "实付金额", "买家实付"],
    "shop": ["店铺名称"],
    "express": ["快递公司"],
    "track": ["快递单号"],
    "seller": ["业务员", "操作员"],
    "customer": ["线下客户", "买家账号"],
}


def _norm(s) -> str:
    return re.sub(r"\s+", "", str(s or "")).lower()


def norm_date(v) -> str:
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    m = re.match(r"(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return s


def read_rows(file: UploadFile):
    try:
        wb = load_workbook(file.file, data_only=True)
    except Exception:
        raise HTTPException(400, "无法读取文件，请上传 .xlsx 格式")
    ws = wb.worksheets[0]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _norm_cell(c) -> str:
    """表头单元格归一化：去首尾空白、去星号（*必填 标记）。"""
    return str(c).strip().replace("*", "").strip() if c is not None else ""


def detect_header(rows, aliases) -> tuple[dict, int]:
    """扫描前 20 行找表头行，返回 {字段: 列索引} 与数据起始行号。"""
    flat = {name for names in aliases.values() for name in names}
    for idx, row in enumerate(rows[:20]):
        cells = {_norm_cell(c) for c in row if _norm_cell(c)}
        hits = cells & flat
        if len(hits) >= 2:
            mapping = {}
            for i, c in enumerate(row):
                c = _norm_cell(c)
                for field, names in aliases.items():
                    if c in names:
                        mapping[field] = i
            return mapping, idx + 1
    return {}, 0


def cell(row, idx, default=""):
    if idx is None or idx >= len(row):
        return default
    v = row[idx]
    if v is None:
        return default
    return str(v).strip()


def to_float(v, default=0.0):
    if v is None:
        return default
    s = str(v).strip()
    if not s:
        return default
    s = s.replace("￥", "").replace("¥", "").replace("元", "").replace("，", ",").replace("；", ";")
    s = s.replace(" ", "").replace("\u00A0", "")
    if s.endswith("元"):
        s = s[:-1]
    s = s.replace(",", "")
    try:
        return float(s)
    except (TypeError, ValueError):
        cleaned = re.sub(r"[^\d\-+\.]+", "", s)
        if cleaned in ("", "-", "+", ".", "-.", "+."):
            return default
        try:
            return float(cleaned)
        except ValueError:
            return default


def infer_base_unit(unit: str) -> str:
    return "克" if unit in WEIGHT_UNITS else "个"


def parse_conversions(text: str, base_unit: str, unit: str) -> dict:
    conv = default_conversions(base_unit)
    if text:
        for part in re.split(r"[;,；，\n]", str(text)):
            part = part.strip()
            if "=" in part:
                u, f = part.split("=", 1)
                u, f = u.strip(), to_float(f)
                if u and f > 0:
                    conv[u] = f
    if unit and unit not in conv:
        if unit in WEIGHT_UNITS:
            conv[unit] = WEIGHT_UNITS[unit]
        else:
            conv[unit] = 1.0  # 计数单位暂定 1:1，需人工核对
    return conv


def parse_pack_items(text: str, db: Session) -> list[dict]:
    """解析 '泡沫箱=1个;泡沫垫=2个'，返回 [{product_id, quantity, unit}]，缺失商品自动跳过。"""
    items = []
    if not text:
        return items
    for part in re.split(r"[;,；，\n]", text):
        part = part.strip()
        if not part:
            continue
        name, _, qty_unit = part.partition("=")
        name = name.strip()
        qty_unit = qty_unit.strip()
        m = re.match(r"^([\d.]+)\s*(.*)$", qty_unit)
        qty = float(m.group(1)) if m else 1.0
        unit = m.group(2).strip() if m and m.group(2).strip() else "个"
        p = resolve_product(db, name)
        if p:
            items.append({"product_id": p.id, "quantity": qty, "unit": unit})
    return items


# ---------------- 模板下载 ----------------
def _xlsx_response(wb, filename: str):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _serve_tpl(filename: str, factory):
    """优先返回 docs/ 目录下的静态模板文件（便于维护），不存在则动态生成。"""
    f = DOCS_DIR / filename
    if f.exists():
        return FileResponse(
            f,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename,
        )
    return _xlsx_response(factory(), filename)


@router.get("/templates/products")
def tpl_products(user: User = Depends(get_current_user)):
    def factory():
        wb = Workbook()
        ws = wb.active
        ws.title = "商品导入"
        ws.append(["商品编码", "商品名称*", "分类", "规格", "基础单位", "单位换算", "默认售价", "打包费", "关联商品清单"])
        ws.append(["", "番茄", "蔬菜", "每个约150克", "克", "克=1;斤=500;公斤=1000;个=150", 0.016, 1, "泡沫箱=1个;泡沫垫=2个"])
        ws.append(["ydj001", "土豆", "蔬菜", "", "斤", "克=1;斤=500;公斤=1000", 2, 0, ""])
        for col, w in zip("ABCDEFGHI", [14, 18, 10, 20, 10, 30, 10, 8, 30]):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"
        return wb
    return _serve_tpl("商品导入模板.xlsx", factory)


@router.get("/templates/inbounds")
def tpl_inbounds(user: User = Depends(get_current_user)):
    def factory():
        wb = Workbook()
        ws = wb.active
        ws.title = "入库导入"
        ws.append(["商品编码或名称*", "进货单位*", "数量*", "单价*", "供应商", "日期*", "操作员", "备注"])
        ws.append(["番茄", "斤", 10, 3, "张三菜行", "2026-08-25", "管理员", ""])
        ws.append(["泡沫箱", "个", 50, 2, "包装厂", "2026-08-25", "", ""])
        for col, w in zip("ABCDEFGH", [20, 12, 10, 10, 16, 14, 12, 16]):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"
        return wb
    return _serve_tpl("入库批量导入模板.xlsx", factory)


@router.get("/templates/outbounds")
def tpl_outbounds(user: User = Depends(get_current_user)):
    def factory():
        wb = Workbook()
        ws = wb.active
        ws.title = "出库导入"
        ws.append(["单号", "日期*", "客户", "商品编码或名称*", "销售单位*", "数量*", "单价*", "打包费", "操作员", "备注"])
        ws.append(["单A001", "2026-08-25", "李四", "番茄", "斤", 3, 8, 1, "", ""])
        ws.append(["单A001", "2026-08-25", "李四", "泡沫垫", "个", 2, 0, 0, "", "同单号自动合并为一单"])
        ws.append(["", "2026-08-25", "王五", "土豆", "公斤", 2, 4, 0, "", "单号留空则每行一单"])
        for col, w in zip("ABCDEFGHIJ", [12, 14, 12, 20, 12, 10, 10, 8, 12, 20]):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"
        return wb
    return _serve_tpl("出库批量导入模板.xlsx", factory)


# ---------------- 商品导入（支持柠檬云模板） ----------------
@router.post("/import/products")
def import_products(file: UploadFile, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    rows = read_rows(file)
    mapping, start = detect_header(rows, PRODUCT_ALIASES)
    if not mapping or "name" not in mapping:
        raise HTTPException(400, "未识别到商品表头（需包含「商品名称」等列），请使用下载的模板或柠檬云商品导入模板")
    success, skipped, failed = 0, 0, []
    created_ids = []
    for i in range(start, len(rows)):
        row = rows[i]
        name = cell(row, mapping.get("name"))
        if not name:
            continue
        try:
            code = cell(row, mapping.get("code"))
            if db.scalar(select(Product).where(Product.name == name)) or (code and db.scalar(select(Product).where(Product.code == code))):
                skipped += 1
                continue
            unit = cell(row, mapping.get("unit"))
            base_unit = cell(row, mapping.get("base_unit")) or infer_base_unit(unit)
            conversions = parse_conversions(cell(row, mapping.get("conversions")), base_unit, unit)
            if base_unit not in conversions:
                conversions[base_unit] = 1.0
            default_unit = unit or base_unit  # 默认出库/展示单位，如 斤
            if default_unit not in conversions:
                conversions[default_unit] = 1.0
            pack_items = parse_pack_items(cell(row, mapping.get("pack_items")), db)
            p = Product(
                code=code,
                name=name,
                category=cell(row, mapping.get("category")),
                spec=cell(row, mapping.get("spec")),
                base_unit=base_unit,
                default_unit=default_unit,
                sale_price=to_float(cell(row, mapping.get("sale_price"))),
                conversions=conversions,
                pack_items=pack_items,
                pack_fee=to_float(cell(row, mapping.get("pack_fee"))),
                is_active=True,
            )
            db.add(p)
            db.flush()
            created_ids.append(p.id)
            success += 1
        except Exception as e:
            failed.append({"row": i + 1, "reason": f"{name}: {e}"})
    db.commit()
    return {
        "ok": True, "created": success, "skipped": skipped,
        "failed": failed, "product_ids": created_ids,
        "failed_count": len(failed),
    }


# ---------------- 入库导入 ----------------
class DraftInbound(BaseModel):
    product_id: int
    product_name: str = ""
    unit: str
    quantity: float
    unit_price: float
    supplier: str = ""
    date: str
    operator: str = ""
    remark: str = ""


class ConfirmInboundIn(BaseModel):
    items: list[DraftInbound] = []


def parse_inbound_draft(file: UploadFile, db: Session, user: User) -> tuple[list[DraftInbound], list[dict]]:
    """解析入库模板 → 草稿入库行（不建单）。"""
    rows = read_rows(file)
    mapping, start = detect_header(rows, INBOUND_ALIASES)
    if not mapping or "product" not in mapping:
        raise HTTPException(400, "未识别到入库表头（需包含「商品」列），请使用下载的入库导入模板")
    items, failed = [], []
    for i in range(start, len(rows)):
        row = rows[i]
        product_key = cell(row, mapping.get("product"))
        if not product_key:
            continue
        product = resolve_product(db, product_key)
        if not product:
            failed.append({"row": i + 1, "reason": f"商品「{product_key}」不存在"})
            continue
        if product.product_type == "order":
            failed.append({"row": i + 1, "reason": f"「{product.name}」是订单商品（小类），请入库其关联的库存商品（大类）"})
            continue
        unit = cell(row, mapping.get("unit"))
        if unit not in (product.conversions or {}):
            failed.append({"row": i + 1, "reason": f"商品「{product.name}」未配置单位「{unit}」"})
            continue
        qty = to_float(cell(row, mapping.get("quantity")), -1)
        price = to_float(cell(row, mapping.get("unit_price")), -1)
        if qty <= 0 or price < 0:
            failed.append({"row": i + 1, "reason": f"数量/单价无效（{product_key}）"})
            continue
        items.append(
            DraftInbound(
                product_id=product.id, product_name=product.name, unit=unit,
                quantity=qty, unit_price=price,
                supplier=cell(row, mapping.get("supplier")),
                operator=cell(row, mapping.get("operator")) or user.name,
                date=norm_date(cell(row, mapping.get("date"))) or datetime.now().strftime("%Y-%m-%d"),
                remark=cell(row, mapping.get("remark")),
            )
        )
    return items, failed


@router.post("/import/inbounds/preview")
def preview_import_inbounds(file: UploadFile, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    items, failed = parse_inbound_draft(file, db, user)
    return {"items": [it.model_dump() for it in items], "failed": failed, "failed_count": len(failed)}


@router.post("/import/inbounds/confirm")
def confirm_import_inbounds(data: ConfirmInboundIn, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    created, failed = 0, []
    for it in data.items:
        if it.quantity <= 0:
            failed.append({"row": it.product_name, "reason": "数量无效"})
            continue
        try:
            create_inbound(
                db,
                {
                    "product_id": it.product_id, "unit": it.unit, "quantity": it.quantity,
                    "unit_price": it.unit_price, "supplier": it.supplier,
                    "operator": it.operator or user.name,
                    "date": it.date, "remark": it.remark,
                },
                operator=user.name,
            )
            db.flush()
            created += 1
        except Exception as e:
            failed.append({"row": it.product_name, "reason": str(e)})
    db.commit()
    return {"ok": True, "created": created, "failed": failed, "failed_count": len(failed)}


@router.post("/import/inbounds")
def import_inbounds(file: UploadFile, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    rows = read_rows(file)
    mapping, start = detect_header(rows, INBOUND_ALIASES)
    if not mapping or "product" not in mapping:
        raise HTTPException(400, "未识别到入库表头（需包含「商品」列），请使用下载的入库导入模板")
    success, failed = 0, []
    for i in range(start, len(rows)):
        row = rows[i]
        product_key = cell(row, mapping.get("product"))
        if not product_key:
            continue
        product = resolve_product(db, product_key)
        if not product:
            failed.append({"row": i + 1, "reason": f"商品「{product_key}」不存在"})
            continue
        unit = cell(row, mapping.get("unit"))
        if unit not in (product.conversions or {}):
            failed.append({"row": i + 1, "reason": f"商品「{product.name}」未配置单位「{unit}」"})
            continue
        qty = to_float(cell(row, mapping.get("quantity")), -1)
        price = to_float(cell(row, mapping.get("unit_price")), -1)
        if qty <= 0 or price < 0:
            failed.append({"row": i + 1, "reason": f"数量/单价无效（{product_key}）"})
            continue
        try:
            create_inbound(
                db,
                {
                    "product_id": product.id, "unit": unit, "quantity": qty,
                    "unit_price": price, "supplier": cell(row, mapping.get("supplier")),
                    "operator": cell(row, mapping.get("operator")) or user.name,
                    "date": norm_date(cell(row, mapping.get("date"))) or datetime.now().strftime("%Y-%m-%d"),
                    "remark": cell(row, mapping.get("remark")),
                },
                operator=user.name,
            )
            success += 1
        except Exception as e:
            failed.append({"row": i + 1, "reason": str(e)})
    db.commit()
    return {"ok": True, "created": success, "failed": failed, "failed_count": len(failed)}


# ---------------- 出库导入 ----------------
class DraftLine(BaseModel):
    product_id: int
    product_name: str = ""
    unit: str
    quantity: float
    price: float
    amount: float = 0.0
    deduct: str = ""  # 扣减说明（订单商品→库存商品）


class DraftOrder(BaseModel):
    doc_no: str = ""
    date: str
    customer: str = ""
    operator: str = ""
    remark: str = ""
    pack_fee: float = 0.0
    lines: list[DraftLine] = []


class ConfirmOutIn(BaseModel):
    orders: list[DraftOrder] = []


def _confirm_orders(db: Session, user: User, orders: list[DraftOrder]) -> dict:
    """按用户确认后的草稿创建出库单（自动结转成本与关联商品）。"""
    created, warnings, failed = 0, [], []
    for o in orders:
        if not o.lines:
            continue
        try:
            rec, warns = create_outbound(
                db,
                {
                    "customer": o.customer, "operator": o.operator or user.name,
                    "date": o.date, "remark": o.remark,
                    "lines": [
                        {"product_id": l.product_id, "unit": l.unit, "quantity": l.quantity, "price": l.price}
                        for l in o.lines
                    ],
                    "pack_lines": [], "pack_fee_total": o.pack_fee or 0,
                },
                operator=user.name,
            )
            db.flush()
            created += 1
            for w in warns:
                warnings.append(f"{rec.code}: {w}")
        except Exception as e:
            failed.append({"doc": o.doc_no, "reason": str(e)})
    db.commit()
    return {"ok": True, "created": created, "failed": failed, "warnings": warnings, "failed_count": len(failed)}


def parse_outbound_draft(file: UploadFile, db: Session, user: User) -> tuple[list[DraftOrder], list[dict]]:
    """解析标准出库模板 → 草稿单（不建单）。"""
    rows = read_rows(file)
    mapping, start = detect_header(rows, OUTBOUND_ALIASES)
    if not mapping or "product" not in mapping:
        raise HTTPException(400, "未识别到出库表头（需包含「商品」列），请使用下载的出库导入模板")
    orders: dict[str, dict] = {}
    failed = []
    for i in range(start, len(rows)):
        row = rows[i]
        product_key = cell(row, mapping.get("product"))
        if not product_key:
            continue
        product = resolve_product(db, product_key)
        if not product:
            failed.append({"row": i + 1, "reason": f"商品「{product_key}」不存在"})
            continue
        unit = cell(row, mapping.get("unit"))
        if unit not in (product.conversions or {}):
            failed.append({"row": i + 1, "reason": f"商品「{product.name}」未配置单位「{unit}」"})
            continue
        qty = to_float(cell(row, mapping.get("quantity")), -1)
        price = to_float(cell(row, mapping.get("unit_price")), 0)
        if qty <= 0:
            failed.append({"row": i + 1, "reason": f"数量无效（{product_key}）"})
            continue
        doc_no = cell(row, mapping.get("doc_no")) or f"_r{i}"
        order = orders.setdefault(
            doc_no,
            {
                "doc_no": doc_no,
                "date": norm_date(cell(row, mapping.get("date"))) or datetime.now().strftime("%Y-%m-%d"),
                "customer": cell(row, mapping.get("customer")),
                "operator": cell(row, mapping.get("operator")) or user.name,
                "remark": cell(row, mapping.get("remark")),
                "pack_fee": 0.0,
                "lines": [],
            },
        )
        order["lines"].append(
            DraftLine(
                product_id=product.id, product_name=product.name, unit=unit,
                quantity=qty, price=price, amount=round(qty * price, 2),
            )
        )
        order["pack_fee"] += to_float(cell(row, mapping.get("pack_fee")))
    return [DraftOrder(**o) for o in orders.values()], failed


def parse_jushuitan_draft(file: UploadFile, db: Session, user: User, statuses: tuple | None = ("已出库",)) -> tuple[list[DraftOrder], list[dict], dict, set]:
    """解析聚水潭出库单 → 草稿单（不建单）。

    statuses 为 None 时接受全部状态（鲜货需求预演算用），默认只取「已出库」（正式导入用）。
    """
    rows = read_rows(file)
    orders, skip = jushuitan_rows(rows, statuses)
    mapping_rows = list(db.execute(select(CodeMapping)).scalars())
    mapping_by_code = {m.external_code: m for m in mapping_rows}

    drafts, failed, unmapped_codes = [], [], set()
    for o in orders:
        lines = []
        for ext_name, qty in parse_jushuitan_name(o["name"]):
            m = mapping_by_code.get(ext_name)
            pid = m.product_id if m else None
            p = db.get(Product, pid) if pid else None
            if not p:
                # 未配置编码关联时，回退按商品名称/编码精确匹配（如「佛手柑中果2个」）
                p = db.scalar(select(Product).where(or_(Product.name == ext_name, Product.code == ext_name)))
            if not p:
                unmapped_codes.add(ext_name)
                continue
            unit, per_item = pick_jst_unit(p, ext_name)
            if unit is None or per_item is None:
                failed.append({"doc": o["doc_no"], "reason": f"「{ext_name}」未配置每件重量换算（如 1个=1000克 或 每件2斤），请在商品管理中补充换算后重试"})
                continue
            if unit not in (p.conversions or {}):
                failed.append({"doc": o["doc_no"], "reason": f"「{ext_name}」换算单位「{unit}」未配置"})
                continue
            # 消耗量 = 件数 × 每件数量（如 1件×2斤=2斤）
            lines.append({"p": p, "ext_name": ext_name, "unit": unit, "qty": round(qty * per_item, 4)})
        if not lines:
            failed.append({"doc": o["doc_no"], "reason": "无已关联商品（未关联编码）"})
            continue

        # 金额按商品默认售价比例分摊卖家实收
        revenue = o["amount"]
        qty_bases, raws = [], []
        for ln in lines:
            qb = unit_to_base(ln["p"], ln["unit"], ln["qty"])
            ln["qty_base"] = qb
            qty_bases.append(qb)
            raws.append(ln["p"].sale_price * qb)
        sum_raw = sum(raws)
        sum_qb = sum(qty_bases)
        draft_lines = []
        for ln, raw, qb in zip(lines, raws, qty_bases):
            if sum_raw > 0:
                amt = revenue * raw / sum_raw
            else:
                amt = revenue * qb / sum_qb if sum_qb > 0 else 0
            amt = round(amt, 2)
            draft_lines.append(
                DraftLine(
                    product_id=ln["p"].id, product_name=ln["p"].name, unit=ln["unit"],
                    quantity=ln["qty"], price=round(amt / ln["qty"], 4) if ln["qty"] else 0,
                    amount=amt, deduct=f"{ln['ext_name']} 每件{fmt_qty(per_item)}{unit}" if per_item else "",
                )
            )
        drafts.append(
            DraftOrder(
                doc_no=o["doc_no"], date=o["date"],
                customer=o["customer"] or o["shop"],
                operator=o["seller"] or user.name,
                remark=f"聚水潭导入 单{o['doc_no']} {o['express']}{o['track']}",
                pack_fee=0.0, lines=draft_lines,
            )
        )
    return drafts, failed, skip, unmapped_codes


@router.post("/import/outbounds/preview")
def preview_import_outbounds(file: UploadFile, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    drafts, failed = parse_outbound_draft(file, db, user)
    return {"orders": [o.model_dump() for o in drafts], "failed": failed, "failed_count": len(failed)}


@router.post("/jushuitan/import/preview")
def preview_import_jushuitan(file: UploadFile, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    drafts, failed, skip, unmapped = parse_jushuitan_draft(file, db, user)
    return {
        "orders": [o.model_dump() for o in drafts],
        "skip": skip, "failed": failed, "failed_count": len(failed),
        "unmapped_codes": sorted(unmapped),
    }


@router.post("/import/outbounds/confirm")
def confirm_import_outbounds(data: ConfirmOutIn, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    return _confirm_orders(db, user, data.orders)


@router.post("/jushuitan/import/confirm")
def confirm_import_jushuitan(data: ConfirmOutIn, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    return _confirm_orders(db, user, data.orders)


@router.post("/import/outbounds")
def import_outbounds(file: UploadFile, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    rows = read_rows(file)
    mapping, start = detect_header(rows, OUTBOUND_ALIASES)
    if not mapping or "product" not in mapping:
        raise HTTPException(400, "未识别到出库表头（需包含「商品」列），请使用下载的出库导入模板")
    orders: dict[str, dict] = {}
    failed = []
    for i in range(start, len(rows)):
        row = rows[i]
        product_key = cell(row, mapping.get("product"))
        if not product_key:
            continue
        product = resolve_product(db, product_key)
        if not product:
            failed.append({"row": i + 1, "reason": f"商品「{product_key}」不存在"})
            continue
        unit = cell(row, mapping.get("unit"))
        if unit not in (product.conversions or {}):
            failed.append({"row": i + 1, "reason": f"商品「{product.name}」未配置单位「{unit}」"})
            continue
        qty = to_float(cell(row, mapping.get("quantity")), -1)
        price = to_float(cell(row, mapping.get("unit_price")), 0)
        if qty <= 0:
            failed.append({"row": i + 1, "reason": f"数量无效（{product_key}）"})
            continue
        doc_no = cell(row, mapping.get("doc_no")) or f"_r{i}"
        order = orders.setdefault(
            doc_no,
            {
                "date": norm_date(cell(row, mapping.get("date"))) or datetime.now().strftime("%Y-%m-%d"),
                "customer": cell(row, mapping.get("customer")),
                "operator": cell(row, mapping.get("operator")) or user.name,
                "remark": cell(row, mapping.get("remark")),
                "lines": [],
                "fee": 0.0,
            },
        )
        order["lines"].append(
            {"product_id": product.id, "unit": unit, "quantity": qty, "price": price}
        )
        order["fee"] += to_float(cell(row, mapping.get("pack_fee")))

    created, warnings = 0, []
    for doc_no, order in orders.items():
        if not order["lines"]:
            continue
        try:
            rec, warns = create_outbound(
                db,
                {
                    "customer": order["customer"], "operator": order["operator"],
                    "date": order["date"], "remark": order["remark"],
                    "lines": order["lines"], "pack_lines": [], "pack_fee_total": order["fee"],
                },
                operator=user.name,
            )
            db.flush()
            created += 1
            for w in warns:
                warnings.append(f"{rec.code}: {w}")
        except Exception as e:
            failed.append({"row": doc_no, "reason": str(e)})
    db.commit()
    return {"ok": True, "created": created, "failed": failed, "warnings": warnings, "failed_count": len(failed)}


# ---------------- 聚水潭：解析 / 关联 / 导入 ----------------
def _normalize_jushuitan_status(value) -> str:
    return re.sub(r"[\s\u3000]+", "", str(value or "")).strip()


def parse_jushuitan_name(name: str) -> list[tuple[str, float]]:
    """将聚水潭导出商品串拆成 (商品名, 数量) 列表，兼容中英文分隔符及序号前缀。"""
    out = []
    s = str(name or "").strip()
    if not s:
        return out
    s = s.replace("，", ",").replace("；", ";").replace("、", ",").replace("\n", ";").replace("\r", ";")
    s = re.sub(r"^\d+\.?\d*\s*\.\s*", "", s)
    for part in re.split(r"[;,]", s):
        part = part.strip()
        if not part:
            continue
        part = re.sub(r"^\d+\.?\d*\s*\.\s*", "", part)
        pm = re.match(r"^(.*?)\*(\d+(?:\.\d+)?)\s*$", part)
        if pm:
            product = pm.group(1).strip()
            qty = float(pm.group(2))
        else:
            product = part
            qty = 1.0
        product = product.strip("()[]{} ")
        if not product:
            continue
        out.append((product, qty))
    return out


# 聚水潭商品名中的单件规格：如 "京鲜生七彩花生2斤" → 每件 2斤
SPEC_RE = re.compile(r"([\d.]+)\s*(斤|公斤|千克|克|g|kg)")
UNIT_ALIAS = {"g": "克", "kg": "千克"}
# 计数单位优先顺序（含订单商品的「单」：1件订单商品 = 1 单）
COUNT_PREF = ["个", "单", "袋", "包", "盒", "箱", "件", "份"]


def parse_jst_spec(name: str) -> tuple[str | None, float | None]:
    """从外部商品名解析每件规格，如 '京鲜生七彩花生2斤' → ('斤', 2.0)。"""
    m = SPEC_RE.search(str(name or ""))
    if m:
        u = UNIT_ALIAS.get(m.group(2), m.group(2))
        return u, float(m.group(1))
    return None, None


def pick_jst_unit(product: Product, ext_name: str) -> tuple[str | None, float | None]:
    """决定聚水潭一行用哪个单位与每件数量（件数→实际数量的倍数关系）。

    优先级：
    1) 商品名内嵌规格（如 每件2斤 → 单位斤、每件数量2）；
    2) 商品已配置的计数单位（个/袋/包…，1件=1个/袋）；
    3) 基础单位本身是计数类；
    否则返回 (None, None)，交由调用方明确报错，避免错误入账。
    """
    conv = product.conversions or {}
    u, qty = parse_jst_spec(ext_name)
    if u and u in conv:
        return u, qty
    if u and product.base_unit == u:
        return u, qty
    for cu in COUNT_PREF:
        if cu in conv:
            return cu, 1.0
    if product.base_unit in COUNT_PREF:
        return product.base_unit, 1.0
    return None, None


def jushuitan_rows(rows, statuses: tuple | None = ("已出库",)) -> list[dict]:
    """解析聚水潭行。statuses 为 None 时接受全部状态（鲜货需求预演算），仅排除 作废/已删除/空。"""
    mapping, start = detect_header(rows, JUSHUITAN_COLS)
    if not mapping or "name" not in mapping:
        raise HTTPException(400, "未识别到聚水潭出库单表头（需包含「出库单号」「商品名称」等列）")
    result, skip = [], {"待出库": 0, "作废": 0, "其他": 0}
    for row in rows[start:]:
        status = _normalize_jushuitan_status(cell(row, mapping.get("status")))
        if statuses is None:
            if status in ("作废", "已删除", ""):
                skip[status if status in skip else "其他"] += 1
                continue
        elif status not in statuses:
            skip[status if status in skip else "其他"] += 1
            continue
        result.append(
            {
                "doc_no": cell(row, mapping.get("doc_no")),
                "date": norm_date(cell(row, mapping.get("date"))),
                "name": cell(row, mapping.get("name")),
                "amount": to_float(cell(row, mapping.get("amount"))),
                "shop": cell(row, mapping.get("shop")),
                "express": cell(row, mapping.get("express")),
                "track": cell(row, mapping.get("track")),
                "seller": cell(row, mapping.get("seller")),
                "customer": cell(row, mapping.get("customer")),
            }
        )
    return result, skip


def auto_suggest(name: str, products: list[Product]) -> tuple[int | None, float]:
    n = _norm(name)
    best, best_score = None, 0.0
    for p in products:
        for cand in (p.name, p.code):
            c = _norm(cand)
            if not c:
                continue
            if c == n:
                return p.id, 1.0
            score = difflib.SequenceMatcher(None, n, c).ratio()
            if score > best_score:
                best_score, best = score, p
    if best and best_score >= 0.35:
        return best.id, round(best_score, 3)
    return None, 0.0


# 自动新增时排除的库存分类（人工/包材/快递不作为关联大类）
_NO_AUTO_STOCK_CATS = ("人工", "包材", "快递")

# 尾部规格剥离：如「新鲜芦笋2斤」「新鲜天麻2斤8-10个」「牛奶芋头3斤30g＋」→ 基础名
_SPEC_TAIL_RE = re.compile(
    r"[\d.]+(?:\s*[－\-–]\s*[\d.]+)?\s*(?:斤|公斤|千克|克|g|kg|个|袋|包|盒|箱|件|份|瓶|支)[+\＋]?$"
)


def _strip_spec(name: str) -> str:
    """剥离商品名尾部的规格，得到基础名（如 新鲜芦笋2斤 → 新鲜芦笋）。"""
    s = str(name or "").strip()
    while True:
        n = _SPEC_TAIL_RE.sub("", s).strip()
        if n == s:
            return s
        s = n


def _match_stock(db: Session, ext_name: str) -> Product | None:
    """从外部商品名自动匹配对应的库存商品（大类，排除 人工/包材/快递）。

    策略：先剥离尾部规格后按基础名精确匹配库存商品名/编码（安全）；
          未命中时再按整名高阈值(0.7)模糊匹配，避免误配（如 西兰苔 ≠ 西兰花）。
    """
    base = _strip_spec(ext_name)
    stocks = [
        p for p in db.execute(select(Product).where(Product.product_type == "stock")).scalars()
        if p.category not in _NO_AUTO_STOCK_CATS
    ]
    for s in stocks:
        if s.name == base or s.code == base:
            return s
    if base == ext_name:
        return None
    pid, score = auto_suggest(base, stocks)
    if pid and score >= 0.7:
        return db.get(Product, pid)
    return None


def _spec_multiplier(stock: Product, ext_name: str) -> float:
    """按外部商品名规格折算到库存默认单位的倍数（1 单订单商品消耗多少库存默认单位）。

    如「新鲜芦笋2斤」→ 2斤=1公斤（库存默认公斤）；「佛手柑大果2个」→ 2个；「新鲜芦笋250g」→ 0.25公斤。
    """
    du = stock.default_unit or stock.base_unit
    conv = stock.conversions or {}
    factor_du = conv.get(du, 1) or 1
    u, qty = parse_jst_spec(ext_name)  # 重量单位：斤/公斤/克
    if u and u in conv:
        return round(qty * conv[u] / factor_du, 6)
    m = re.search(r"([\d.]+)\s*(个|袋|包|盒|箱|件|份|瓶|支)", ext_name)  # 计数单位
    if m:
        cu = m.group(2)
        if du == cu:
            return round(float(m.group(1)), 6)
        f = conv.get(cu)
        if f:
            return round(float(m.group(1)) * f / factor_du, 6)
    return 1.0


def _auto_create_order(db: Session, ext_name: str, stock: Product | None = None) -> Product:
    """为外部商品自动新增订单商品（小类）。stock 为空则仅新增、不关联库存（待后续维护）。"""
    existing = db.scalar(select(Product).where(Product.name == ext_name))
    if existing:
        return existing
    if stock:
        mult = _spec_multiplier(stock, ext_name)
        spec = f"每单约 {fmt_qty(mult)}{stock.default_unit or stock.base_unit}"
        sid = stock.id
        cat = stock.category
    else:
        mult = 1.0
        spec = "未关联库存，待匹配"
        sid = None
        cat = "待分类"
    p = Product(
        code=ext_name,
        name=ext_name,
        category=cat,
        product_type="order",
        base_unit="单",
        default_unit="单",
        spec=spec,
        conversions={"单": 1},
        stock_product_id=sid,
        multiplier=mult,
        is_active=True,
    )
    db.add(p)
    db.flush()
    return p


@router.post("/jushuitan/parse")
def parse_jushuitan(file: UploadFile, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    # 关联页用于「发现商品并自动新增品类」，与出库状态无关：接受全部状态（仅排除作废）
    rows = read_rows(file)
    orders, skip = jushuitan_rows(rows, statuses=None)
    mapping_rows = list(db.execute(select(CodeMapping)).scalars())
    mapping_by_code = {m.external_code: m for m in mapping_rows}

    codes: dict[str, dict] = {}
    for o in orders:
        for ext_name, qty in parse_jushuitan_name(o["name"]):
            c = codes.setdefault(ext_name, {"external_code": ext_name, "count": 0})
            c["count"] += 1

    auto_created = 0
    for c in codes.values():
        su, sq = parse_jst_spec(c["external_code"])
        c["spec"] = f"{fmt_qty(sq)} {su}/件" if su and sq else ""
        m = mapping_by_code.get(c["external_code"])
        if m and m.product_id:
            # 已有映射：如指向订单商品（小类），修正为指向其关联的库存商品（大类）
            c["stock_product_name"] = ""
            tp = db.get(Product, m.product_id)
            if tp and tp.product_type == "order" and tp.stock_product_id:
                sp = db.get(Product, tp.stock_product_id)
                if sp and sp.product_type == "stock":
                    m.product_id = sp.id
                    c["stock_product_name"] = sp.name
            sp = db.get(Product, m.product_id)
            c["product_id"] = m.product_id
            c["product_name"] = tp.name if tp else ""
            c["stock_product_id"] = m.product_id if sp and sp.product_type == "stock" else None
            c["stock_product_name"] = c["stock_product_name"] or (sp.name if sp and sp.product_type == "stock" else "")
            c["score"] = m.auto_score
            c["status"] = "已关联"
            continue
        # 订单商品（小类）是否已存在：存在则不新增
        existed = db.scalar(select(Product).where(Product.name == c["external_code"]))
        if existed:
            order_p = existed
            c["status"] = "已存在"
        else:
            # 不存在：自动新增订单商品品类，并尝试与库存商品（大类）关键词关联
            stock = _match_stock(db, c["external_code"])
            order_p = _auto_create_order(db, c["external_code"], stock)
            auto_created += 1
            c["status"] = "自动新增"
        # 已存在但未关联库存的，补关键词关联
        if c["status"] == "已存在" and not order_p.stock_product_id:
            stock = _match_stock(db, c["external_code"])
            if stock:
                order_p.stock_product_id = stock.id
                order_p.multiplier = _spec_multiplier(stock, c["external_code"])
                order_p.spec = f"每单约 {fmt_qty(order_p.multiplier)}{stock.default_unit or stock.base_unit}"
        # 编码关联统一指向库存商品（大类）；无匹配库存则不关联（留待人工）
        sp = db.get(Product, order_p.stock_product_id) if order_p.stock_product_id else None
        if sp and sp.product_type == "stock":
            _upsert_mapping(db, "jushuitan", c["external_code"], sp.id)
            c["product_id"] = sp.id
            c["stock_product_id"] = sp.id
            c["stock_product_name"] = sp.name
        else:
            c["product_id"] = order_p.id
            c["stock_product_id"] = None
            c["stock_product_name"] = ""
        c["product_name"] = order_p.name
        c["multiplier"] = order_p.multiplier
        c["score"] = 1.0
    db.commit()
    return {
        "total_orders": len(orders),
        "skip": skip,
        "auto_created": auto_created,
        "codes": sorted(codes.values(), key=lambda x: -x["count"]),
    }


@router.get("/mappings")
def list_mappings(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    rows = db.execute(select(CodeMapping).order_by(CodeMapping.id)).scalars()
    return [
        {
            "id": m.id, "source": m.source, "external_code": m.external_code,
            "external_name": m.external_name, "product_id": m.product_id,
            "product_name": m.product.name if m.product else "",
            "auto_score": m.auto_score,
        }
        for m in rows
    ]


class MappingIn(BaseModel):
    source: str = "jushuitan"
    external_code: str
    product_id: int | None = None


class MappingBulkIn(BaseModel):
    source: str = "jushuitan"
    items: list[MappingIn]


def _upsert_mapping(db: Session, source: str, external_code: str, product_id: int | None):
    m = db.scalar(
        select(CodeMapping).where(CodeMapping.source == source, CodeMapping.external_code == external_code)
    )
    if m:
        m.product_id = product_id
        m.updated_at = datetime.now()
    else:
        db.add(CodeMapping(source=source, external_code=external_code, product_id=product_id))
    db.flush()


@router.post("/mappings")
def set_mapping(data: MappingIn, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    _upsert_mapping(db, data.source, data.external_code, data.product_id)
    db.commit()
    return {"ok": True}


@router.post("/mappings/bulk")
def bulk_mappings(data: MappingBulkIn, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    for item in data.items:
        _upsert_mapping(db, data.source, item.external_code, item.product_id)
    db.commit()
    return {"ok": True, "saved": len(data.items)}


@router.post("/mappings/auto")
def auto_mappings(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    # 只自动关联到库存商品（大类），且用基础名精确/高阈值匹配，避免旧版模糊误配（如 西兰苔≠西兰花）
    rows = list(db.execute(select(CodeMapping).where(CodeMapping.product_id.is_(None))).scalars())
    matched = 0
    for m in rows:
        stock = _match_stock(db, m.external_code)
        if stock:
            m.product_id = stock.id
            m.auto_score = 1.0
            m.updated_at = datetime.now()
            matched += 1
    db.commit()
    return {"ok": True, "matched": matched, "total": len(rows)}


@router.delete("/mappings")
def clear_mappings(source: str = "jushuitan", db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    for m in db.execute(select(CodeMapping).where(CodeMapping.source == source)).scalars():
        db.delete(m)
    db.commit()
    return {"ok": True}


@router.delete("/mappings/{mid}")
def delete_mapping(mid: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    m = db.get(CodeMapping, mid)
    if not m:
        raise HTTPException(404, "关联记录不存在")
    db.delete(m)
    db.commit()
    return {"ok": True}


@router.post("/jushuitan/import")
def import_jushuitan(file: UploadFile, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    rows = read_rows(file)
    orders, skip = jushuitan_rows(rows)
    mapping_rows = list(db.execute(select(CodeMapping)).scalars())
    mapping_by_code = {m.external_code: m for m in mapping_rows}

    created, warnings, failed = 0, [], []
    unmapped_codes = set()
    for o in orders:
        lines = []
        for ext_name, qty in parse_jushuitan_name(o["name"]):
            m = mapping_by_code.get(ext_name)
            pid = m.product_id if m else None
            p = db.get(Product, pid) if pid else None
            if not p:
                # 未配置编码关联时，回退按商品名称/编码精确匹配（如「佛手柑中果2个」）
                p = db.scalar(select(Product).where(or_(Product.name == ext_name, Product.code == ext_name)))
            if not p:
                unmapped_codes.add(ext_name)
                continue
            unit, per_item = pick_jst_unit(p, ext_name)
            if unit is None or per_item is None:
                failed.append({"doc": o["doc_no"], "reason": f"「{ext_name}」未配置每件重量换算（如 1个=1000克 或 每件2斤），请在商品管理中补充换算后重试"})
                continue
            if unit not in (p.conversions or {}):
                failed.append({"doc": o["doc_no"], "reason": f"「{ext_name}」换算单位「{unit}」未配置"})
                continue
            # 消耗量 = 件数 × 每件数量（如 1件×2斤=2斤）
            lines.append({"product": p, "ext_name": ext_name, "unit": unit, "qty": round(qty * per_item, 4)})
        if not lines:
            failed.append({"doc": o["doc_no"], "reason": "无已关联商品（未关联编码）"})
            continue

        # 金额按商品默认售价比例分摊卖家实收
        revenue = o["amount"]
        qty_bases, raws = [], []
        for ln in lines:
            qb = unit_to_base(ln["product"], ln["unit"], ln["qty"])
            ln["qty_base"] = qb
            qty_bases.append(qb)
            raws.append(ln["product"].sale_price * qb)
        sum_raw = sum(raws)
        sum_qb = sum(qty_bases)
        for ln, raw, qb in zip(lines, raws, qty_bases):
            if sum_raw > 0:
                amt = revenue * raw / sum_raw
            else:
                amt = revenue * qb / sum_qb if sum_qb > 0 else 0
            ln["amount"] = round(amt, 2)
            ln["price"] = round(amt / ln["qty"], 4) if ln["qty"] else 0
        try:
            rec, warns = create_outbound(
                db,
                {
                    "customer": o["customer"] or o["shop"],
                    "operator": o["seller"] or user.name,
                    "date": o["date"],
                    "remark": f"聚水潭导入 单{o['doc_no']} {o['express']}{o['track']}",
                    "lines": [
                        {"product_id": ln["product"].id, "unit": ln["unit"], "quantity": ln["qty"], "price": ln["price"]}
                        for ln in lines
                    ],
                    "pack_lines": [],
                    "pack_fee_total": None,
                },
                operator=user.name,
            )
            db.flush()
            created += 1
            for w in warns:
                warnings.append(f"{rec.code}: {w}")
        except Exception as e:
            failed.append({"doc": o["doc_no"], "reason": str(e)})

    db.commit()
    return {
        "ok": True, "created": created, "skip": skip,
        "failed": failed, "warnings": warnings,
        "unmapped_codes": sorted(unmapped_codes),
        "failed_count": len(failed),
    }
