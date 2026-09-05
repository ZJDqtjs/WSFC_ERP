"""
货物需求清单统计脚本
功能：
1. 解析"商品信息"字段，展开多商品订单为单商品行
2. 将商品编码按"大类"合并（如 佛手柑中果1个/2个/4个 → 佛手柑中果）
3. 统计每个大类的总需求量（单位：个、盒、件、斤、g）
4. 混合单位（如新鲜芦笋既有g又有斤）统一换算为"斤"（1斤=500g）
说明：
- 数据来源："商品信息"字段，格式 `[序号.]商品编码*数量[单位]`，多个商品用逗号分隔
- 需求 = sum(数量) × 每包规格  （数量由"商品信息"解析得到）
- 大类提取：剥离编码中所有 "数字+单位(个/盒/斤/g/装)" 和 "净重"
- 单位优先级：个 > 盒 > 斤 > g
- 未识别单位的商品默认统计为"件"
"""
import re
import os
import sys
import argparse
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def parse_info(info):
    """解析"商品信息"字段，返回 [(商品编码, 数量), ...]

    格式: `[序号.]商品编码*数量[单位]`，多个商品用逗号(中/英)分隔
    特殊处理：`*N盒` 格式时，将"N盒"保留在商品编码中（数量设为1），
              这样 extract_spec 能优先识别到"盒"单位，且不同规格分别显示。
    示例:
      '1.新鲜香蕈菌250g*1'                                 -> [('新鲜香蕈菌250g', 1)]
      '2.佛手柑中果1个*2'                                   -> [('佛手柑中果1个', 2)]
      '3.京鲜生五指毛桃250g*1,京鲜生鸡骨草250g*1,石斛鲜条250g*1' -> [(...), (...), (...)]
      '4.包浆豆腐300g*1盒'                                 -> [('包浆豆腐300g1盒', 1)]
      '5.包浆豆腐300g*2盒'                                 -> [('包浆豆腐300g2盒', 1)]
    无 `*数量` 后缀的视为数量1。
    """
    items = []
    for part in re.split(r'[,，]', str(info)):
        part = part.strip()
        if not part:
            continue
        # 先匹配 *数字盒：将"数字盒"保留在商品编码中，数量设为1
        m = re.match(r'^(?:\d+\.)?(.+?)\*(\d+)盒$', part)
        if m:
            items.append((m.group(1).strip() + m.group(2) + "盒", 1))
            continue
        # 再匹配 *数字[件]?
        m = re.match(r'^(?:\d+\.)?(.+?)\*(\d+)(?:件)?$', part)
        if m:
            items.append((m.group(1).strip(), int(m.group(2))))
        else:
            name = re.sub(r'^\d+\.', '', part).strip()
            if name:
                items.append((name, 1))
    return items


def extract_spec(code: str):
    """从商品编码中提取每包规格数量与单位。优先级：个 → 盒 → 斤 → g。未识别单位默认为"件"。"""
    code = str(code)
    m = re.search(r"(\d+(?:\.\d+)?)个$", code)
    if m:
        return float(m.group(1)), "个"
    m = re.search(r"(\d+(?:\.\d+)?)盒", code)
    if m:
        return float(m.group(1)), "盒"
    m = re.search(r"(\d+(?:\.\d+)?)斤", code)
    if m:
        return float(m.group(1)), "斤"
    m = re.search(r"(\d+(?:\.\d+)?)g\+?$", code)
    if m:
        return float(m.group(1)), "g"
    return 1.0, "件"


def extract_category(code: str):
    """从商品编码中提取大类名称（剥离所有规格数量与单位信息）"""
    code = str(code)
    # 先删 g 相关（g+、g），再删 斤 相关（斤装、斤），然后删 盒、个，最后删 净重
    code = re.sub(r"(\d+(?:\.\d+)?)g\+?", "", code)
    code = re.sub(r"(\d+(?:\.\d+)?)斤装?", "", code)
    code = re.sub(r"(\d+(?:\.\d+)?)盒", "", code)
    code = re.sub(r"(\d+(?:\.\d+)?)个", "", code)
    code = re.sub(r"净重", "", code)
    code = code.replace("*", "").strip()
    return code


def to_jin(amount, unit):
    """将需求量统一换算为斤（用于跨单位合并）。盒、件、个保持不变。"""
    if unit in ("个", "盒", "件"):
        return amount, unit
    if unit == "斤":
        return amount, "斤"
    if unit == "g":
        return amount / 500, "斤"
    return amount, unit


def _extract_fruit_weight(code: str):
    """从商品编码提取单品果重量(g)，如 100g+, 250g+, 500g+, 120g+, 180g+"""
    code = str(code)
    m = re.search(r"(\d+(?:\.\d+)?)g\+", code)
    if m:
        return float(m.group(1))
    # 甜龙笋8斤500g+ 这种
    m = re.search(r"(\d+(?:\.\d+)?)g", code)
    if m:
        return float(m.group(1))
    return None


# 品类映射表：内部大类 -> 图片品类名
# 键: (内部大类, 单品果重量g) -> (图片品类名, 单果kg)
_CATEGORY_MAP = {
    # 佛手柑（个单位，按个数统计）
    ("佛手柑果王", None): ("佛手柑2斤+", 1.0),      # 2斤/果 = 1kg
    ("佛手柑特大果", None): ("佛手柑800+", 0.8),     # 800g/果
    ("佛手柑大果", None): ("佛手柑600+", 0.6),      # 600g/果
    ("佛手柑中果", None): ("佛手柑400+", 0.4),      # 400g/果
    
    # 根茎类
    ("黑紫土豆", None): ("黑紫土豆", None),
    ("黑紫土豆", 80.0): ("七彩土豆", None),          # 80g单果 → 七彩土豆
    # 甜龙笋（按单果重量拆分300+/500+）
    ("甜龙笋", 250.0): ("甜笋300+", None),
    ("甜龙笋", 500.0): ("甜笋500+", None),
    # 洋蓟（按单果重量拆分100+/150+/200+）
    ("新鲜洋蓟", 100.0): ("洋蓟100+", None),
    ("新鲜洋蓟", 150.0): ("洋蓟150+", None),
    ("新鲜洋蓟", 200.0): ("洋蓟200+", None),
    # 西红柿/马蹄番茄（按单果重量）
    ("西红柿", 120.0): ("西红柿120+", None),
    ("马蹄番茄", 180.0): ("马蹄番茄180+", None),
    # 其他蔬菜/水果
    ("人参果中果", None): ("人参果", None),
    ("人参果大果", None): ("人参果", None),
    ("人参果特大果", None): ("人参果", None),
    ("小米辣", None): ("小米辣", None),
    ("小花糯玉米", None): ("小花糯玉米", None),
    ("拇指小玉米", None): ("拇指小玉米", None),
    ("新鲜玉米笋", None): ("玉米笋", None),
    ("新鲜芦笋", None): ("芦笋", None),
    ("玉米段", None): ("玉米段", None),
    ("玉米粒", None): ("玉米粒", None),
    ("紫拇指玉米", None): ("紫拇指玉米", None),
    ("紫皮独头蒜中果", None): ("紫皮独头蒜", None),
    ("紫皮独头蒜大果", None): ("紫皮独头蒜", None),
    ("京鲜生五指毛桃", None): ("五指毛桃", None),
    ("京鲜生鸡骨草", None): ("鸡骨草", None),
    ("新鲜香蕈菌", None): ("香蕈", None),
    ("水果玉米", None): ("水果玉米", None),
    ("石斛鲜条", None): ("石斛", None),
    ("水果西芹", None): ("西芹", None),
}


def _map_to_order_item(row):
    """将一行数据映射到订货单的品类名和单果重量"""
    cat = row["大类"]
    fw = row.get("单品果重量g", None)
    unit = row["单位"]

    # 优先匹配有具体单果重量的
    key = (cat, fw)
    if key in _CATEGORY_MAP:
        name, fruit_kg = _CATEGORY_MAP[key]
        return name, fruit_kg

    # 回退：只用大类名匹配
    key = (cat, None)
    if key in _CATEGORY_MAP:
        name, fruit_kg = _CATEGORY_MAP[key]
        return name, fruit_kg

    # 未匹配则用大类名
    return cat, None


# 订货单图片板式：品类顺序 + 出品率
_ORDER_LAYOUT = [
    ("七彩土豆", 0.85),
    ("芦笋", 0.95),
    ("黑紫土豆", 0.9),
    ("小米辣", 0.92),
    ("西红柿120+", 0.95),
    ("马蹄番茄180+", 0.9),
    ("小芋头", 0.9),
    ("小芋头100+", 0.9),
    ("西芹", 0.9),
    ("佛手柑2斤+", 1),
    ("佛手柑800+", 1),
    ("佛手柑600+", 1),
    ("香蕈", 0.99),
    ("玉米笋", 0.8),
    ("甜笋300+", 0.9),
    ("甜笋500+", 0.9),
    ("石斛", 1),
    ("洋蓟100+", 0.95),
    ("洋蓟150+", 0.95),
    ("洋蓟200+", 0.95),
]


def _build_order_sheet(order_agg, date_str):
    """构建订货单数据行列表"""
    today = date_str
    # 建立映射：品类名 -> 订单数量(kg或个)
    qty_map = dict(zip(order_agg["订货单品类"], order_agg["订单公斤"]))

    items = []
    for cat_name, yield_rate in _ORDER_LAYOUT:
        order_val = qty_map.get(cat_name, 0)
        is_piece = cat_name.startswith("佛手柑")

        if order_val == 0:
            order_str = ""
        elif is_piece:
            order_str = int(order_val)
        elif float(order_val).is_integer():
            order_str = int(order_val)
        else:
            order_str = round(order_val, 2)

        items.append({
            "日期": today,
            "品类": cat_name,
            "库存": "",
            "订单": order_str,
            "出品率": yield_rate,
            "备注": "",
        })

    # 添加板式中没有但我们有的品类（放在末尾）
    extra_items = set(qty_map.keys()) - set(cat for cat, _ in _ORDER_LAYOUT)
    for cat in sorted(extra_items):
        order_val = qty_map.get(cat, 0)
        if order_val == 0:
            continue
        is_piece = cat.startswith("佛手柑")
        if is_piece:
            order_str = int(order_val)
        else:
            order_str = int(order_val) if float(order_val).is_integer() else round(order_val, 2)
        items.append({
            "日期": today,
            "品类": cat,
            "库存": "",
            "订单": order_str,
            "出品率": "",
            "备注": "",
        })

    return items


def run(input_file, out_dir=None, debug=False):
    """GUI/命令行统一入口。

    input_file : 需求清单 .xlsx 路径
    out_dir    : 输出目录（缺省为输入文件所在目录）
    debug      : 调试模式（只打印解析详情，不生成文件）
    返回输出文件路径（调试模式返回 None）。
    """
    argv = [str(input_file)]
    argv.append("--select" if debug else "--")
    if out_dir:
        argv.extend(["--out-dir", str(out_dir)])
    argv = [a for a in argv if a != "--"]
    main(argv)


def main(argv=None):
    parser = argparse.ArgumentParser(description="货物需求清单统计脚本")
    parser.add_argument("input", nargs="?", help="输入的需求清单 .xlsx 文件路径")
    parser.add_argument("--select", action="store_true", help="调试模式：显示解析详情，不生成Excel文件")
    parser.add_argument("--out-dir", default=None, help="输出目录（缺省为输入文件所在目录）")
    args = parser.parse_args(argv)

    if not args.input:
        print("错误：请提供需求清单 .xlsx 文件路径！")
        print("用法：python stat_demand.py <需求清单.xlsx> [--select]")
        print("      python stat_demand.py --select  (可在脚本内指定默认文件)")
        sys.exit(1)

    SRC = args.input
    if not os.path.isfile(SRC):
        print(f"错误：找不到文件 {SRC}")
        sys.exit(1)

    # 生成输出路径
    base_dir = os.path.dirname(SRC)
    DST = os.path.join(args.out_dir or base_dir, "需求统计_统计核对表.xlsx")
    debug_mode = args.select

    if debug_mode:
        print(f"[调试模式] 正在处理: {SRC}")
    else:
        print(f"正在处理: {SRC}")
        print(f"输出文件: {DST}")
    print("-" * 40)
    
    df_raw = pd.read_excel(SRC)

    # 检查"状态"字段，分离作废单（不参与统计）
    if "状态" in df_raw.columns:
        mask_void = df_raw["状态"].astype(str).str.strip() == "作废"
        df_valid_raw = df_raw[~mask_void].copy()
        df_void_raw = df_raw[mask_void].copy()
        void_count = len(df_void_raw)
        if void_count > 0:
            print(f"检测到 {void_count} 条作废单，已从统计中排除")
        else:
            print("未检测到作废单")
    else:
        df_valid_raw = df_raw
        df_void_raw = pd.DataFrame()
        print("未找到'状态'列，全部数据参与统计")
    print("-" * 40)

    # 解析"商品信息"字段，将一行多商品展开为多行单商品
    # 每个展开行携带：商品编码、数量、出库日期、原始变体（用于排除多单商品）
    expanded_rows = []
    for _, r in df_valid_raw.iterrows():
        info = r.get("商品信息", "")
        out_date = r.get("出库日期", None)
        for part in re.split(r'[,，]', str(info)):
            part = part.strip()
            if not part:
                continue
            item_str = re.sub(r'^\d+\.', '', part).strip()
            if not item_str:
                continue
            for name, qty in parse_info(part):
                expanded_rows.append({
                    "商品编码": name,
                    "数量": qty,
                    "出库日期": out_date,
                    "原始变体": item_str,
                })
    df = pd.DataFrame(expanded_rows)

    # 统计每种商品变体出现在多少个订单中（用于sheet4多单统计）
    # 只统计带 `*数字` 格式的商品（如 香菇干货500g*2），不带*的不统计
    variant_order_counts = {}
    for _, r in df_valid_raw.iterrows():
        info = r.get("商品信息", "")
        seen = set()
        for part in re.split(r'[,，]', str(info)):
            part = part.strip()
            if not part:
                continue
            item_str = re.sub(r'^\d+\.', '', part).strip()
            if item_str and re.search(r'\*\d+', item_str):
                seen.add(item_str)
        for item_str in seen:
            variant_order_counts[item_str] = variant_order_counts.get(item_str, 0) + 1

    # 特殊单变体集合：所有带 `*数字` 格式的商品条目
    # （如京鲜生鸡骨草250g*2）全部放入sheet4统计，sheet1/sheet2不再包含
    special_variants = set(variant_order_counts.keys())

    # 排除所有特殊单（已在sheet4中单独统计），后续sheet1/sheet2不再重复统计
    if special_variants:
        before = len(df)
        df = df[~df["原始变体"].isin(special_variants)].copy()
        excluded = before - len(df)
        print(f"已排除 {excluded} 个特殊单商品行（在sheet4中单独统计）")
    else:
        excluded = 0

    print(f"有效订单 {len(df_valid_raw)} 行，展开为 {len(df)} 个商品行，"
          f"共 {df['商品编码'].nunique()} 个唯一商品编码，"
          f"特殊单 {len(special_variants)} 种")
    print("-" * 40)

    # 提取每包规格与单位
    df["每包规格"], df["单位"] = zip(*df["商品编码"].apply(extract_spec))

    # 提取大类名称
    df["大类"] = df["商品编码"].apply(extract_category)

    # 校验：未识别单位的商品编码已默认归为"件"
    defaulted = df[df["单位"] == "件"]
    if not defaulted.empty:
        print(f"提示：以下 {defaulted['商品编码'].nunique()} 个商品编码无明确单位，已默认统计为「件」：")
        for code in defaulted["商品编码"].unique():
            print(f"  - {code}")
        print()

    # 计算每行需求量 = 数量 × 每包规格
    # 数量来自"商品信息"字段解析，每行已是单商品
    df["需求量_raw"] = df["数量"] * df["每包规格"]

    # 按大类聚合
    agg_funcs = {
        "商品编码": ("商品编码", "first"),
        "单位": ("单位", lambda x: x.mode().iloc[0] if not x.mode().empty else "斤"),
        "需求量_raw": ("需求量_raw", "sum"),
    }

    grouped = df.groupby("大类", as_index=False).agg(**agg_funcs)

    # 处理混合单位：如果某大类有多种单位，全部换算为斤（个、盒、件除外）
    def finalize(row):
        cat = row["大类"]
        unit = row["单位"]
        total = row["需求量_raw"]
        # 检查该大类是否存在混合单位
        sub = df[df["大类"] == cat]
        units_in_cat = sub["单位"].unique()

        if unit in ("个", "盒", "件"):
            # 个、盒、件 单位的大类通常不会混合，直接使用
            return total, unit
        elif len(units_in_cat) > 1:
            # 混合单位，全部换算为斤
            converted = 0.0
            for _, r in sub.iterrows():
                v, u = to_jin(r["数量"] * r["每包规格"], r["单位"])
                converted += v
            return round(converted, 2), "斤"
        else:
            return total, unit

    grouped["需求量"], grouped["最终单位"] = zip(*grouped.apply(finalize, axis=1))

    def fmt_qty(v):
        return int(v) if float(v).is_integer() else round(v, 2)

    # 大类 -> 最终单位 映射（用于小类排序对齐大类）
    cat_unit_map = dict(zip(grouped["大类"], grouped["最终单位"]))
    unit_order = {"个": 0, "盒": 1, "件": 2, "斤": 3, "g": 4}

    # === 规格明细（小类）===
    spec_detail = (
        df.groupby(["大类", "商品编码", "单位", "每包规格"], as_index=False)
        .agg(包裹总数=("数量", "sum"))
    )

    # 计算每个小类的需求量 = 包裹总数 × 每包规格
    spec_detail["该规格需求量"] = spec_detail["包裹总数"] * spec_detail["每包规格"]

    # === 关键修正：将小类需求量统一换算到大类的最终单位 ===
    # 对于混合单位的大类，小类中g单位的需求要换算为斤（1斤=500g）
    def convert_to_final(row):
        cat = row["大类"]
        spec_unit = row["单位"]
        qty = row["该规格需求量"]
        final_unit = cat_unit_map.get(cat, spec_unit)

        if spec_unit == final_unit:
            return qty, spec_unit
        elif spec_unit == "g" and final_unit == "斤":
            return qty / 500, "斤"
        elif spec_unit in ("个", "盒", "件"):
            return qty, spec_unit
        else:
            return qty, spec_unit

    spec_detail["换算需求量"], spec_detail["显示单位"] = zip(
        *spec_detail.apply(convert_to_final, axis=1)
    )

    def fmt_qty(v):
        return int(v) if float(v).is_integer() else round(v, 2)

    spec_detail["换算需求量"] = spec_detail["换算需求量"].apply(fmt_qty)
    spec_detail["每包规格"] = spec_detail["每包规格"].apply(fmt_qty)
    spec_detail["包裹总数"] = spec_detail["包裹总数"].astype(int)

    # 显示需求量用换算后的值 + 最终单位
    spec_detail["规格需求(带单位)"] = (
        spec_detail["换算需求量"].astype(str) + spec_detail["显示单位"]
    )

    spec_detail["_u"] = spec_detail["大类"].map(cat_unit_map).map(unit_order).fillna(99)
    spec_detail = spec_detail.sort_values(by=["_u", "大类", "商品编码"]).drop(columns="_u").reset_index(drop=True)

    # === 大类汇总 ===
    grouped["需求量"] = grouped["需求量"].apply(fmt_qty)
    grouped["需求量(带单位)"] = grouped["需求量"].astype(str) + grouped["最终单位"]
    grouped["_u"] = grouped["最终单位"].map(unit_order).fillna(99)
    grouped = grouped.sort_values(by=["_u", "大类"]).drop(
        columns=["_u", "需求量_raw", "单位"]
    ).reset_index(drop=True)

    # === 单表：小类明细 + 末尾追加 大类 / 总需求量 两列，方便直接查看 ===
    # 大类 -> 总需求量(带单位) 映射
    cat_total_map = dict(zip(grouped["大类"], grouped["需求量(带单位)"]))

    wb_headers = [
        "大类名称", "规格编码", "单位", "每包规格", "包裹总数",
        "需求量(带单位)", "大类", "总需求量",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "需求统计"

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    # 末尾"大类 / 总需求量"两列高亮，便于直接查看
    summary_fill = PatternFill("solid", fgColor="FCE4D6")
    summary_font = Font(bold=True)

    # 表头
    for c, h in enumerate(wb_headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # 数据行：每个小类一行，末尾附 大类 + 总需求量
    row = 2
    for _, r in spec_detail.iterrows():
        cat = r["大类"]
        vals = [
            cat, r["商品编码"],
            r["显示单位"],  # 使用换算后的最终单位，而非原始单位
            r["每包规格"],
            r["包裹总数"],
            r["规格需求(带单位)"],
            cat, cat_total_map.get(cat, ""),
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = border
            h = wb_headers[c - 1]
            # 先统一居中，合并后再调整
            cell.alignment = center
            if h in ("大类", "总需求量"):
                cell.fill = summary_fill
                cell.font = summary_font
        row += 1

    # 合并单元格：同一大类的 "大类名称"(列1)、"大类"(列7)、"总需求量"(列8) 三列
    merge_cols = [1, 7, 8]

    # 找出每个大类的行号范围（Excel行号从2开始）
    ranges = []
    cat_col = spec_detail["大类"].tolist()
    n = len(cat_col)
    i = 0
    while i < n:
        j = i
        while j < n and cat_col[j] == cat_col[i]:
            j += 1
        # 大类 cat_col[i] 占据 Excel 行 2+i 到 2+j-1
        if j - i > 1:
            ranges.append((2 + i, 2 + j - 1))
        i = j

    # 执行合并
    for start_r, end_r in ranges:
        for col in merge_cols:
            ws.merge_cells(
                start_row=start_r, start_column=col,
                end_row=end_r, end_column=col,
            )
            cell = ws.cell(row=start_r, column=col)
            cell.alignment = center
            cell.border = border
            # 为合并区域的每行设置左右边框（避免边框丢失）
            for r in range(start_r, end_r + 1):
                for c in [1, 7, 8]:
                    ws.cell(row=r, column=c).border = border

    # 列宽
    widths = [18, 22, 8, 12, 12, 16, 18, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 行高
    ws.row_dimensions[1].height = 28
    for r in range(2, row):
        ws.row_dimensions[r].height = 22

    # === 添加订货单sheet ===
    # df已在上方排除多单商品，直接使用
    df_order = df

    # 从df_order按商品编码细分，提取单品果重量用于映射
    df_order["单品果重量g"] = df_order["商品编码"].apply(_extract_fruit_weight)
    df_order["单果公斤"] = df_order["单品果重量g"] / 1000.0

    # 按映射键分组统计（映射键 = 图片中的品类名）
    df_order["订货单品类"], df_order["订货单单果kg"] = zip(*df_order.apply(_map_to_order_item, axis=1))

    # 先逐行换算为公斤（使用映射后的单果重量）
    # 佛手柑品类按个数统计，不换算为公斤
    def row_to_kg(row):
        unit = row["单位"]
        qty = row["需求量_raw"]
        order_cat = str(row["订货单品类"])
        if order_cat.startswith("佛手柑"):
            return qty
        if unit == "斤":
            return qty / 2
        elif unit == "g":
            return qty / 1000
        elif unit in ("个", "盒", "件"):
            fruit_kg = row["订货单单果kg"]
            if fruit_kg and fruit_kg > 0:
                return qty * fruit_kg
            return qty
        else:
            return qty

    df_order["行需求kg"] = df_order.apply(row_to_kg, axis=1)

    # 按订货单品类聚合（已换算为kg）
    order_agg = (
        df_order.groupby("订货单品类", as_index=False)
        .agg(
            订单公斤=("行需求kg", "sum"),
        )
    )

    # 从出库日期字段提取日期（只取年月日）
    date_str = pd.to_datetime(df_order["出库日期"].iloc[0]).strftime("%Y/%m/%d")

    # 构建订货单数据
    order_items = _build_order_sheet(order_agg, date_str)

    # === 调试模式：输出详细解析信息，不生成Excel ===
    if debug_mode:
        pd.set_option("display.unicode.east_asian_width", True)
        pd.set_option("display.width", 220)
        pd.set_option("display.max_columns", None)

        print("\n" + "=" * 60)
        print("【调试模式 - 解析详情】")
        print("=" * 60)

        print("\n--- 1. 商品信息展开结果 ---")
        print(f"原始订单 {len(df_raw)} 行 → 展开 {len(df)} 个商品行，{df['商品编码'].nunique()} 个唯一编码")

        print("\n--- 2. 品类提取结果（前20条）---")
        cat_show = df[['商品编码', '每包规格', '单位', '大类', '数量', '需求量_raw']].head(20)
        print(cat_show.to_string(index=False))

        print("\n--- 3. 大类汇总 ---")
        grouped_show = grouped[['大类', '需求量', '最终单位']].copy()
        grouped_show['需求量(带单位)'] = grouped['需求量(带单位)']
        print(grouped_show.to_string(index=False))

        print("\n--- 4. 订货单品类映射 ---")
        map_show = df_order[['商品编码', '大类', '单品果重量g', '订货单品类', '订货单单果kg', '行需求kg']].drop_duplicates()
        print(map_show.to_string(index=False))

        print("\n--- 5. 订货单汇总 ---")
        print(order_agg.to_string(index=False))

        print("\n--- 6. 订货单明细 ---")
        for item in order_items:
            flag = " [个]" if str(item["品类"]).startswith("佛手柑") else ""
            print(f"  {item['品类']:<14s}{flag}  订单={item['订单']}  出品率={item['出品率']}")

        print("\n--- 7. 规格明细（前30条）---")
        show = spec_detail[['大类', '商品编码', '显示单位', '每包规格', '包裹总数', '规格需求(带单位)']].copy()
        show.columns = ['大类', '商品编码', '单位', '每包规格', '包裹总数', '规格需求(带单位)']
        print(show.head(30).to_string(index=False))

        print(f"\n--- 8. 特殊单统计（带*号的商品，前30条）---")
        sorted_variants = sorted(variant_order_counts.items(), key=lambda x: (-x[1], x[0]))
        for variant, count in sorted_variants[:30]:
            print(f"  {variant:<35s}  {count}单")
        print(f"  ... 共 {len(variant_order_counts)} 种特殊单")

        print(f"\n[调试模式] 共 {len(spec_detail)} 个规格小类，{len(grouped)} 个大类")
        print("[调试模式] 未生成Excel文件")
        return

    # === 正常模式：生成Excel ===
    # 写入订货单sheet
    ws2 = wb.create_sheet(title="订货单")

    # 标题行：A1:G1 合并 "现采鲜货"
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    title_cell = ws2.cell(row=1, column=1, value="现采鲜货")
    title_cell.alignment = center
    title_cell.font = Font(bold=True, size=14)
    title_cell.border = border
    for c in range(1, 8):
        ws2.cell(row=1, column=c).border = border

    # 表头行（第2行）
    headers = ["日期", "品类", "库存", "订单", "出品率", "需求", "备注"]
    for c, h in enumerate(headers, start=1):
        cell = ws2.cell(row=2, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # 数据行（从第3行开始）
    start_data_row = 3
    for i, item in enumerate(order_items):
        r = start_data_row + i
        is_piece = str(item["品类"]).startswith("佛手柑")

        ws2.cell(row=r, column=1, value=item["日期"]).border = border
        ws2.cell(row=r, column=1).alignment = center

        ws2.cell(row=r, column=2, value=item["品类"]).border = border
        ws2.cell(row=r, column=2).alignment = center

        ws2.cell(row=r, column=3, value=item["库存"]).border = border
        ws2.cell(row=r, column=3).alignment = center

        ws2.cell(row=r, column=4, value=item["订单"]).border = border
        ws2.cell(row=r, column=4).alignment = center

        yield_cell = ws2.cell(row=r, column=5, value=item["出品率"])
        yield_cell.border = border
        yield_cell.alignment = center
        if item["出品率"]:
            yield_cell.number_format = "0.00"

        # 需求 = (订单 - 库存) / 出品率，保留两位小数
        demand_cell = ws2.cell(row=r, column=6)
        if item["出品率"] and item["出品率"] > 0:
            demand_cell.value = f"=IF(E{r}=0,0,ROUND((D{r}-C{r})/E{r},2))"
        else:
            demand_cell.value = ""
        demand_cell.border = border
        demand_cell.alignment = center
        demand_cell.number_format = "0.00" if not is_piece else "0"

        ws2.cell(row=r, column=7, value=item["备注"]).border = border
        ws2.cell(row=r, column=7).alignment = left

    # 列宽
    widths2 = [12, 18, 10, 10, 10, 10, 20]
    for i, w in enumerate(widths2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # 行高
    ws2.row_dimensions[1].height = 32
    ws2.row_dimensions[2].height = 28
    for r in range(start_data_row, start_data_row + len(order_items)):
        ws2.row_dimensions[r].height = 22

    # 冻结前两行（标题+表头）
    ws2.freeze_panes = "A3"

    # === 添加作废单sheet ===
    if not df_void_raw.empty:
        ws3 = wb.create_sheet(title="作废单")

        void_headers = ["出库单号", "出库日期", "状态", "店铺名称", "商品信息", "买家账号", "卖家备注"]
        # 只保留存在的列
        void_cols = [c for c in void_headers if c in df_void_raw.columns]
        void_data = df_void_raw[void_cols].copy()

        # 表头
        void_fill = PatternFill("solid", fgColor="C00000")
        void_font = Font(bold=True, color="FFFFFF", size=11)
        for c, h in enumerate(void_cols, start=1):
            cell = ws3.cell(row=1, column=c, value=h)
            cell.fill = void_fill
            cell.font = void_font
            cell.alignment = center
            cell.border = border

        # 数据行
        for r_idx, (_, r) in enumerate(void_data.iterrows(), start=2):
            for c_idx, col in enumerate(void_cols, start=1):
                val = r[col]
                if pd.isna(val):
                    val = ""
                cell = ws3.cell(row=r_idx, column=c_idx, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # 列宽
        void_widths = [18, 14, 10, 16, 40, 18, 24]
        for i, w in enumerate(void_widths, start=1):
            if i <= len(void_cols):
                ws3.column_dimensions[get_column_letter(i)].width = w

        # 行高
        ws3.row_dimensions[1].height = 28
        for r in range(2, len(void_data) + 2):
            ws3.row_dimensions[r].height = 22

        ws3.freeze_panes = "A2"

    # === 添加多单统计sheet ===
    # 统计每种带 `*数字` 格式的特殊单（全部显示，按单数降序排列）
    ws4 = wb.create_sheet(title="多单统计")

    ws4_headers = ["商品信息", "单数"]
    for c, h in enumerate(ws4_headers, start=1):
        cell = ws4.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # 显示所有特殊单，按单数降序排列
    sorted_variants = sorted(variant_order_counts.items(), key=lambda x: (-x[1], x[0]))
    for r_idx, (variant, count) in enumerate(sorted_variants, start=2):
        cell_a = ws4.cell(row=r_idx, column=1, value=variant)
        cell_a.border = border
        cell_a.alignment = left
        cell_b = ws4.cell(row=r_idx, column=2, value=count)
        cell_b.border = border
        cell_b.alignment = center

    # 列宽
    ws4.column_dimensions['A'].width = 35
    ws4.column_dimensions['B'].width = 10

    # 行高
    ws4.row_dimensions[1].height = 28
    for r in range(2, len(sorted_variants) + 2):
        ws4.row_dimensions[r].height = 22

    ws4.freeze_panes = "A2"

    wb.save(DST)

    # 控制台输出
    pd.set_option("display.unicode.east_asian_width", True)
    pd.set_option("display.width", 220)
    pd.set_option("display.max_columns", None)

    show = spec_detail[[
        "大类", "商品编码", "显示单位", "每包规格", "包裹总数", "规格需求(带单位)"
    ]].copy()
    show.columns = ["大类", "商品编码", "单位", "每包规格", "包裹总数", "规格需求(带单位)"]
    show["总需求量"] = show["大类"].map(cat_total_map)
    print(f"共 {len(spec_detail)} 个规格小类，{len(grouped)} 个大类")
    print(f"结果已保存至: {DST}\n")
    print(show.to_string(index=False))
    print("\n处理完成！")


if __name__ == "__main__":
    main()
